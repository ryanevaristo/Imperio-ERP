import openpyxl
from openpyxl.styles import Font

def exportar_tabela_para_xlsx():
    # Criar um novo workbook e worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notificações Médicas"

    # Cabeçalhos
    headers = [
        "Tipo de Notificação", "Listas", "Medicamentos", "Abrangência", 
        "Cor da Notificação", "Quantidade Máxima por receita e período de tratamento", 
        "Quantidade máxima por receita", "Validade da Receita", 
        "Talão da notificação impresso às expensas de"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Dados
    dados = [
        ["Notificação de Receita 'A'", "A1; A2; A3", "Entorpecentes", "Em todo o território nacional", "Amarela", "5 ampolas e demais formas farmacêuticas tratamento p/30 dias", "1 medicamento ou substância", "30 dias", "Autoridade Sanitária – Talão c/ 20 folhas"],
        ["Notificação de Receita 'B'", "B1", "Psicotrópicos", "Na Unidade Federada onde for concedida a Numeração", "Azul", "5 ampolas e demais formas farmacêuticas tratamento p/60 dias", "1 medicamento ou substância", "30 dias", "Profissional"],
        ["Notificação de Receita 'B2'", "B2", "Psicotrópicos Anorexígenos", "Na Unidade Federada onde for concedida a Numeração", "Azul", "Tratamento para no máximo 30 dias", "1 medicamento ou substância", "30 dias", "Profissional"],
        ["Notificação de Receita 'Retinoides'", "C2", "Retinoides Sistêmico", "Em todo o território nacional", "Branca", "5 ampolas e demais formas farmacêuticas tratamento p/30 dias", "1 medicamento ou substância", "30 dias", "Profissional"],
        ["Notificação de Receita Talidomida", "C3", "Imunossupressores", "Tratamento para no máximo 30 dias", "1 medicamento ou substância", "15 dias", "Serviços Públicos de Saúde"],
        ["Receita de Controle Especial", "C1", "Controle Especial", "5 ampolas e demais formas farmacêuticas tratamento p/60 dias", "3 medicamentos ou substâncias", "30 dias", "Profissional"],
        ["Receita de Controle Especial", "C5", "Anabolizantes", "5 ampolas e demais formas farmacêuticas tratamento p/60 dias", "3 medicamentos ou substâncias", "30 dias", "Profissional"],
        ["Receita de Controle Especial", "C4", "Anti-retrovirais", "Todo o Território Nacional", "Branca", "5 ampolas e demais formas farmacêuticas tratamento p/30 dias", "3 medicamentos ou substâncias", "30 dias", "Programa DST/AIDS"],
        ["Receita de Controle Especial ou Comum em 02(duas) Vias", "A1; A2; A3; B1", "Adendos das Listas", "5 ampolas e demais formas farmacêuticas tratamento p/30 dias", "3 medicamentos ou substâncias", "30 dias", "Profissional"],
        ["Receita de Controle Especial ou Comum em 02(duas) Vias", "C1; B1", "Antiparkinsonianos e Anticonvulsivantes", "5 ampolas e demais formas farmacêuticas tratamento p/180 dias", "3 medicamentos ou substâncias", "30 dias", "Profissional"]
    ]

    for row_num, row_data in enumerate(dados, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # Salvar o arquivo
    nome_arquivo = "Notificacoes_Medicas.xlsx"
    wb.save(nome_arquivo)
    return nome_arquivo

# Chamar a função para exportar os dados
nome_arquivo = exportar_tabela_para_xlsx()