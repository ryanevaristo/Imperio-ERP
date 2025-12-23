from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.template import loader
from .models import Movimentacao, Produtos, EstoqueCategoria, Notificacao, MovimentacaoItem
from django.shortcuts import redirect
from django.urls import reverse
from django.contrib import messages
from rolepermissions.decorators import has_role_decorator
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator


from django.template.loader import render_to_string
from weasyprint import HTML
import openpyxl
import uuid

from produto.models import Empreendimento, Lote, Quadra



# Create your views here.

@login_required(login_url='/login/')
@has_role_decorator(["Administrador", "Gerente","estoquista"])
def home_estoque(request):
    # Otimizado: Usa select_related para categoria e order_by para consistência
    produtos = Produtos.objects.select_related('categoria').filter(
        empresa=request.user.empresa
    ).order_by('-created_at')

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        produtos = produtos.filter(created_at__range=[start_date, end_date])

    #Filtra por pesquisar se o parâmetro estiver presente
    pesquisar = request.GET.get('pesquisar')
    if pesquisar:
        produtos = produtos.filter(produto__icontains=pesquisar)

    #paginator
    paginator = Paginator(produtos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Otimizado: Só busca empreendimentos se necessário
    empreendimentos = None
    if not start_date or not end_date:
        empreendimentos = Empreendimento.objects.filter(empresa=request.user.empresa)

    return render(request, 'estoque/home.html', {
        'page_obj': page_obj,
        'start_date': start_date,
        'end_date': end_date,
        'pesquisar': pesquisar,
        'empreendimentos': empreendimentos
    })

def detalhes_produto(request, id):
    produto = Produtos.objects.get(id=id, empresa=request.user.empresa)
    return render(request, 'estoque/detalhes_produto.html', {'produto': produto})

@login_required(login_url='/login/')
@has_role_decorator(["Administrador", "Gerente","estoquista"])
def cadastrar_produto(request):
    cadastrar_categoria = EstoqueCategoria.objects.filter(empresa=request.user.empresa)
    if request.method == 'GET':
        return render(request, 'estoque/cadastrar_editar_produto.html', {'categorias': cadastrar_categoria})
    
    elif request.method == 'POST' and request.POST.get('nome_categoria') =='':
        produto = request.POST.get('produto')
        qtd = request.POST.get('qtd')
        qtd_min = request.POST.get('qtd_min')
        custo = request.POST.get('custo')

        categoria = request.POST.get('categoria')
        # Aqui você deve salvar os dados do produto no banco de dados
        # calculo_margem = (float(venda) - float(custo)) / float(custo) * 100
        # margem = round(calculo_margem, 2)

        produtos = Produtos(
            produto=produto,
            qtd=qtd,
            qtd_min=qtd_min,
            custo=custo,
            categoria=EstoqueCategoria.objects.get(id=categoria)
        )
        produtos.save()
        messages.success(request, 'Produto cadastrado com sucesso!')
        return redirect(reverse('estoque:home_estoque'))
    else:
        nome_categoria = request.POST.get('nome_categoria')
        if EstoqueCategoria.objects.filter(nome_categoria=nome_categoria).exists():
            messages.error(request,'Categoria já cadastrada', extra_tags='danger')
            return redirect("financeiro:cadastrar_despesas")
        
        categorias = EstoqueCategoria(nome_categoria=nome_categoria)
        categorias.save()
        print(request.path_info)
        return HttpResponseRedirect(request.path_info)
        


@login_required(login_url='/login/')
@has_role_decorator(["Administrador", "Gerente","estoquista"])
def editar_produto(request, id):
    produto = Produtos.objects.get(id=id, empresa=request.user.empresa)
    cadastrar_categoria = EstoqueCategoria.objects.filter(empresa=request.user.empresa)
    if request.method == 'GET':
        return render(request, 'estoque/cadastrar_editar_produto.html', {'produto': produto, 'categorias': cadastrar_categoria})
    
    elif request.method == 'POST':
        produto.produto = request.POST.get('produto')
        produto.qtd = request.POST.get('qtd')
        produto.qtd_min = request.POST.get('qtd_min')
        produto.custo = request.POST.get('custo')
        categoria_id = request.POST.get('categoria')
        produto.categoria = EstoqueCategoria.objects.get(id=categoria_id)
        produto.save()
        messages.success(request, 'Produto editado com sucesso!')
        return redirect(reverse('estoque:home_estoque'))


@login_required(login_url='/login/')
@has_role_decorator(["Administrador", "Gerente","estoquista"])
def deletar_produto(request, id):
    produto = Produtos.objects.get(id=id)
    produto.delete()
    messages.success(request, 'Produto deletado com sucesso!')
    return redirect(reverse('estoque:home_estoque'))


@login_required(login_url='/auth/login/')
@has_role_decorator(["gerente", "administrador"])
def cadastrar_categorias(request):
    if request.method == "GET":
        cadastrar_categorias = EstoqueCategoria.objects.filter(empresa=request.user.empresa)
        return render(request, 'cadastrar_categoria_estoque.html', {'categorias': cadastrar_categorias})
    elif request.method == "POST":
        descricao = request.POST.get('nome_categoria')
        if EstoqueCategoria.objects.filter(descricao=descricao, empresa=request.user.empresa).exists():
            messages.error(request, 'Categoria já existe!', extra_tags='danger')
            return redirect(reverse('estoque:cadastrar_categorias'))
        print(descricao)
        # Aqui você deve salvar os dados da categoria no banco de dados
        categoria = EstoqueCategoria(descricao=descricao, empresa=request.user.empresa)
        categoria.save()
        return redirect('/estoque/')
    

def importar_estoque_excel(request):
    if request.method == 'POST' and request.FILES['file']:
        arquivo = request.FILES['file']
        wb = openpyxl.load_workbook(arquivo)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            produto_nome, qtd,qtd_min, custo, venda, categoria_nome = row

            categoria, created = EstoqueCategoria.objects.get_or_create(nome_categoria=categoria_nome)

            produto, created = Produtos.objects.get_or_create(
                produto=produto_nome,
                defaults={'qtd': qtd, 'qtd_min':qtd_min, 'custo': custo, 'venda': venda, 'categoria': categoria}
            )
            if not created:
                produto.qtd += qtd
                produto.qtd_min = qtd_min
                produto.custo = custo
                produto.venda = venda
                produto.categoria = categoria
                produto.save()

        messages.success(request, 'Produtos importados com sucesso!')
        return redirect(reverse('estoque:home_estoque'))

    return render(request, 'estoque/importar_entrada.html')    




@login_required(login_url='/login/')
@has_role_decorator(["Administrador", "Gerente","estoquista"])
def registrar_movimentacao(request):
    produto_id = request.GET.get("produto_id")
    produtos = Produtos.objects.filter(empresa=request.user.empresa)
    empreendimentos = Empreendimento.objects.filter(empresa=request.user.empresa)

    if request.method == "POST":
        tipo = request.POST.get("tipo_movimentacao")
        motivo = request.POST.get("motivo")
        empreendimento_id = request.POST.get("empreendimento")
        quadra_id = request.POST.get("quadra")
        lote_id = request.POST.get("lote")

        movimentacao = Movimentacao.objects.create(
            tipo=tipo,
            motivo=motivo,
            empreendimento_id=empreendimento_id,
            quadra_id=quadra_id,
            lote_id=lote_id,
            empresa=request.user.empresa
        )

        # Se veio produto_id → movimentação individual
        if produto_id:
            produtos_ids = [produto_id]
        else:
            produtos_ids = request.POST.getlist("produtos")

        for pid in produtos_ids:
            qtd = int(request.POST.get(f"quantidade_{pid}", 0))
            if qtd > 0:
                produto_obj = Produtos.objects.get(id=pid)

                if tipo == 'Entrada':
                    produto_obj.qtd += qtd
                elif tipo == 'Saida':
                    produto_obj.qtd -= qtd
                elif tipo == 'Devolucao':
                    produto_obj.qtd += qtd

                produto_obj.save()

                MovimentacaoItem.objects.create(
                    movimentacao=movimentacao,
                    produto=produto_obj,
                    qtd=qtd,
                    empresa=request.user.empresa
                )

        messages.success(request, "Movimentação registrada com sucesso!")
        return redirect("estoque:home_estoque")

    produtos_previos = Produtos.objects.all()[:5]

    return render(request, "estoque/movimentacao/pdv_movimentacao.html", {
        "produtos": produtos,
        "empreendimentos": empreendimentos,
        "produtos_previos": produtos_previos,
        "produto_id": request.GET.get("produto_id")
    })


@login_required(login_url='/login/')
@has_role_decorator(["Administrador", "Gerente","estoquista"])
def buscar_produtos(request):
    termo = request.GET.get("q", "")
    produtos = Produtos.objects.filter(produto__icontains=termo)[:10]  # limita a 10 resultados
    resultados = [{"id": p.id, "nome": p.produto, "qtd": p.qtd} for p in produtos]
    return JsonResponse(resultados, safe=False)


@login_required(login_url='/login/')
@has_role_decorator(["Administrador", "Gerente","estoquista"])
def historico_todas_movimentacoes(request):
    # Busca todas as movimentações ordenadas da mais recente para a mais antiga
    movimentacoes_query = Movimentacao.objects.select_related(
        'empreendimento',
        'quadra',
        'lote'
    ).prefetch_related(
        'itens__produto'
    ).filter(
        empresa=request.user.empresa
    ).order_by('-created_at')
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date and end_date:
        movimentacoes_query = movimentacoes_query.filter(created_at__range=[start_date, end_date])

    pesquisar = request.GET.get('pesquisar')
    if pesquisar:
        movimentacoes_query = movimentacoes_query.filter(itens__produto__produto__icontains=pesquisar).distinct()

    # Agrupa as movimentações com seus itens
    movimentacoes_agrupadas = {}
    for movimentacao in movimentacoes_query:
        itens = movimentacao.itens.all()
        # Se houver filtro de pesquisa, filtra os itens
        if pesquisar:
            itens = itens.filter(produto__produto__icontains=pesquisar)
        
        if itens.exists():  # Só adiciona se tiver itens
            movimentacoes_agrupadas[str(movimentacao.id)] = {
                'movimentacao': movimentacao,
                'itens': list(itens)
            }

    return render(request, 'estoque/historico_movimentacoes.html', {
        'movimentacoes_agrupadas': movimentacoes_agrupadas,
        'start_date': start_date,
        'end_date': end_date,
        'pesquisar': pesquisar
    })

def get_quadras(request, empreendimento_id):
    quadras = Quadra.objects.filter(empreendimento_id=empreendimento_id)
    data = [{'id': str(q.id), 'nome': q.nome} for q in quadras]
    return JsonResponse({'quadras': data})


def get_lotes(request, quadra_id):
    lotes = Lote.objects.filter(quadra_id=quadra_id)
    data = [{'id': str(l.id), 'numero': l.numero} for l in lotes]
    return JsonResponse({'lotes':data}, safe=False)



@login_required(login_url='/login/')
@has_role_decorator(["Administrador", "Gerente","estoquista"])
def exportar_estoque_xls(request):
    # Otimizado: Usa select_related para categoria e only() para campos necessários
    produtos = Produtos.objects.select_related('categoria').filter(
        empresa=request.user.empresa
    ).only(
        'id', 'produto', 'qtd', 'qtd_min', 'custo', 'categoria__nome_categoria'
    ).order_by('produto')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_estoque.xlsx"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Estoque'

    # Cabeçalhos
    headers = ['ID', 'Produto', 'Quantidade', 'Quantidade Mínima', 'Custo', 'Categoria']
    sheet.append(headers)

    # Dados dos produtos - otimizado para evitar queries adicionais
    for produto in produtos:
        row = [
            str(produto.id) if produto.id else "",
            str(produto.produto) if produto.produto else "",
            str(produto.qtd) if produto.qtd is not None else "0",
            str(produto.qtd_min) if produto.qtd_min is not None else "0",
            str(produto.custo) if produto.custo is not None else "0",
            str(produto.categoria.nome_categoria) if produto.categoria else ""
        ]
        sheet.append(row)

    workbook.save(response)
    return response

@login_required(login_url='/login/')
@has_role_decorator(["Administrador", "Gerente", "estoquista"])
def exportar_estoque_pdf(request):
    # Otimizado: Usa select_related para categoria e only() para campos necessários
    produtos = Produtos.objects.select_related('categoria').filter(
        empresa=request.user.empresa
    ).only(
        'produto', 'qtd', 'qtd_min', 'custo', 'categoria__nome_categoria'
    ).order_by('produto')
    
    html_string = render_to_string('estoque/relatorio_estoque.html', {
        'produtos': produtos,
        'empresa': request.user.empresa
    })
    pdf_file = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=relatorio_estoque.pdf'
    return response

@login_required(login_url='/login/')
@has_role_decorator(["Administrador", "Gerente","estoquista"])
def listar_notificacoes(request):
    notificacoes = Notificacao.objects.filter(visualizado=False).order_by('-data_criacao')
    return render(request, 'estoque/notificacoes.html', {'notificacoes': notificacoes,'notificacoes_count': notificacoes.count()})



@login_required(login_url='/login/')
@has_role_decorator(["Administrador", "Gerente","estoquista"])
def marca_vizualizado(request, id):
    try:
        notificacao = Notificacao.objects.get(id=id)
        notificacao.visualizado = True
        notificacao.save()
        messages.success(request, 'Notificação marcada como visualizada!')
    except Notificacao.DoesNotExist:
        messages.error(request, 'Notificação não encontrada!', extra_tags='danger')
    return redirect(request.META.get('HTTP_REFERER', 'estoque:notificacoes'))



